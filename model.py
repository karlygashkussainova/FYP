import xlsxwriter
import openpyxl
from pathlib import Path
import numpy as np
import math
import statistics
import cv2
import PIL
from PIL import Image
import matplotlib 
from matplotlib import image
from matplotlib import pyplot
from numpy import asarray
import scipy
from scipy.stats import skew, kurtosis
from skimage import io
import skimage.measure    
import skimage.feature
import xlsxwriter
import pandas as pd
import seaborn as sns



# Create File
outWorkbook = xlsxwriter.Workbook("out.xlsx")
outSheet = outWorkbook.add_worksheet()

# Write Headers
outSheet.write("A1", "Mean1")
outSheet.write("B1", "Variance1")
outSheet.write("C1", "Skewness1")
outSheet.write("D1", "Kurtosis1")
outSheet.write("E1", "Contrast1")
outSheet.write("F1", "Entropy1")
outSheet.write("G1", "Energy1")
outSheet.write("H1", "Homo1")
outSheet.write("I1", "Corre1")
outSheet.write("J1", "Mean2")
outSheet.write("K1", "Variance2")
outSheet.write("L1", "Skewness2")
outSheet.write("M1", "Kurtosis2")
outSheet.write("N1", "Contrast2")
outSheet.write("O1", "Entropy2")
outSheet.write("P1", "Energy2")
outSheet.write("Q1", "Homo2")
outSheet.write("R1", "Corre2")
outSheet.write("S1", "Mean3")
outSheet.write("T1", "Variance3")
outSheet.write("U1", "Skewness3")
outSheet.write("V1", "Kurtosis3")
outSheet.write("W1", "Contrast3")
outSheet.write("X1", "Entropy3")
outSheet.write("Y1", "Energy3")
outSheet.write("Z1", "Homo13")
outSheet.write("AA1", "Corre3")
 
outWorkbook.close()

import glob
Image_list = []
for filename in glob.glob('C:/Users/karlygash.kussainova/Desktop/senior/FYP/finalProject/Images/*.png'): #assuming gif
    im=Image.open(filename)
    Image_list.append(im)
import glob
Image_list = []
for filename in glob.glob('C:/Users/karlygash.kussainova/Desktop/senior/FYP/finalProject/Images/*.png'): #assuming gif
    im=Image.open(filename)
    Image_list.append(im)
        
for i in range(0, len(Image_list)):
    Image = Image_list[i]  
    image = asarray(Image) 
    R = image[:,:,0]
    G = image[:,:,1]
    B = image[:,:,2] 
    
    
    # Histogram Equalization for Red Channel
    R = asarray(R) 
    # put pixels in a 1D array by flattening out img array
    flat = R.flatten()
    # Histogram Equalization
    # show the histogram
    pyplot.figure(1)
    pyplot.hist(flat, bins=10)

    # create our own histogram function
    def get_histogram1(R, bins):
        # array with size of bins, set to zeros
        histogram1 = np.zeros(bins)
# loop through pixels and sum up counts of pixels
        for pixel in R:
            histogram1[pixel] += 1
    
# return our final result
        return histogram1
# execute our histogram function
    hist = get_histogram1(flat, 256)
    # create our cumulative sum function
    def cumsum(a1):
        a1 = iter(a1)
        b1 = [next(a1)]
        for i in a1:
            b1.append(b1[-1] + i)
        return np.array(b1)
# execute the fn
    cs1 = cumsum(hist)

# display the result
    pyplot.figure(2)
    pyplot.plot(cs1)

# numerator & denomenator
    nj1 = (cs1 - cs1.min()) * 255
    N1 = cs1.max() - cs1.min()

# re-normalize the cumsum
    cs1 = nj1 / N1

# cast it back to uint8 since we can't use floating point values in images
# pyplot.figure(1)
    cs1 = cs1.astype('uint8')

    pyplot.figure(3)
    pyplot.plot(cs1)

# get the value from cumulative sum for every index in flat, and set that as img_new
    img_new1 = cs1[flat]

# put array back into original shape since we flattened it
    img_new1 = np.reshape(img_new1, R.shape)

# set up side-by-side image display
# fig = pyplot.figure(3)
# fig.set_figheight(15)
# fig.set_figwidth(15)

# fig.add_subplot(1,2,1)
# pyplot.figure(4)
# pyplot.imshow(R, cmap='gray')

# display the new image
# fig.add_subplot(1,2,2)
    pyplot.figure(4)
    pyplot.imshow(img_new1, cmap='gray')

# pyplot.show(block=True)

    R = img_new1


  # Histogram Equalization for Red Channel
    G = asarray(G) 
    # put pixels in a 1D array by flattening out img array
    flat = G.flatten()
    # Histogram Equalization
    # show the histogram
    pyplot.figure(1)
    pyplot.hist(flat, bins=10)

    # create our own histogram function
    def get_histogram2(G, bins):
        # array with size of bins, set to zeros
        histogram2 = np.zeros(bins)
# loop through pixels and sum up counts of pixels
        for pixel in G:
            histogram2[pixel] += 1
    
# return our final result
        return histogram2
# execute our histogram function
    hist = get_histogram2(flat, 256)
    # create our cumulative sum function
    def cumsum(a2):
        a2 = iter(a2)
        b2 = [next(a2)]
        for i in a2:
            b2.append(b2[-1] + i)
        return np.array(b2)
# execute the fn
    cs2 = cumsum(hist)

# display the result
    pyplot.figure(2)
    pyplot.plot(cs2)

# numerator & denomenator
    nj2 = (cs2 - cs2.min()) * 255
    N2 = cs2.max() - cs2.min()

# re-normalize the cumsum
    cs2 = nj2 / N2

# cast it back to uint8 since we can't use floating point values in images
# pyplot.figure(1)
    cs2 = cs2.astype('uint8')

    pyplot.figure(3)
    pyplot.plot(cs2)

# get the value from cumulative sum for every index in flat, and set that as img_new
    img_new2 = cs2[flat]

# put array back into original shape since we flattened it
    img_new2 = np.reshape(img_new2, G.shape)

# set up side-by-side image display
# fig = pyplot.figure(3)
# fig.set_figheight(15)
# fig.set_figwidth(15)

# fig.add_subplot(1,2,1)
# pyplot.figure(4)
# pyplot.imshow(R, cmap='gray')

# display the new image
# fig.add_subplot(1,2,2)
    pyplot.figure(4)
    pyplot.imshow(img_new2, cmap='gray')

# pyplot.show(block=True)

    G = img_new2
    
    
    
    
    # Histogram Equalization for Red Channel
    B = asarray(B) 
    # put pixels in a 1D array by flattening out img array
    flat = B.flatten()
    # Histogram Equalization
    # show the histogram
    pyplot.figure(1)
    pyplot.hist(flat, bins=10)

    # create our own histogram function
    def get_histogram3(B, bins):
        # array with size of bins, set to zeros
        histogram3 = np.zeros(bins)
# loop through pixels and sum up counts of pixels
        for pixel in B:
            histogram3[pixel] += 1
    
# return our final result
        return histogram3
# execute our histogram function
    hist = get_histogram3(flat, 256)
    # create our cumulative sum function
    def cumsum(a3):
        a3 = iter(a3)
        b3 = [next(a3)]
        for i in a3:
            b3.append(b3[-1] + i)
        return np.array(b3)
# execute the fn
    cs3 = cumsum(hist)

# display the result
    pyplot.figure(3)
    pyplot.plot(cs3)

# numerator & denomenator
    nj3 = (cs3 - cs3.min()) * 255
    N3 = cs3.max() - cs3.min()

# re-normalize the cumsum
    cs3 = nj3 / N3

# cast it back to uint8 since we can't use floating point values in images
# pyplot.figure(1)
    cs3 = cs3.astype('uint8')

    pyplot.figure(3)
    pyplot.plot(cs3)

# get the value from cumulative sum for every index in flat, and set that as img_new
    img_new3 = cs3[flat]

# put array back into original shape since we flattened it
    img_new3 = np.reshape(img_new3, B.shape)

# set up side-by-side image display
# fig = pyplot.figure(3)
# fig.set_figheight(15)
# fig.set_figwidth(15)

# fig.add_subplot(1,2,1)
# pyplot.figure(4)
# pyplot.imshow(R, cmap='gray')

# display the new image
# fig.add_subplot(1,2,2)
    pyplot.figure(4)
    pyplot.imshow(img_new3, cmap='gray')

# pyplot.show(block=True)

    B = img_new3
    
       
    # Feature Selection from Red Channel
    print ("Calucation of First Four Statistical Four Moments for Red Channel")
    Mean1 = np.mean(img_new1);
    print (Mean1)
    Variance1 = np.var(R)
    Variance1 = math.sqrt(Variance1)
    print(Variance1)
    Skewness1=skew(R.reshape(-1))
    print (Skewness1)
    Kurtosis1=kurtosis(R.reshape(-1))
    print (Kurtosis1)
    entropy1 = skimage.measure.shannon_entropy(R)
    print (entropy1)
    R = skimage.img_as_ubyte(R)
    g1 = skimage.feature.greycomatrix(R, [1], [0], levels=256, symmetric=False, normed=True)
    Cont1 =skimage.feature.greycoprops(g1, 'contrast')[0][0]
    print (Cont1)
    Energ1 =skimage.feature.greycoprops(g1, 'energy')[0][0]
    print (Energ1)
    Homo1=skimage.feature.greycoprops(g1, 'homogeneity')[0][0]
    print (Homo1)
    Corre1=skimage.feature.greycoprops(g1, 'correlation')[0][0]
    print (Corre1)
    
    
    # Feature Selection from Green Channel
    print ("Calucation of First Four Statistical Four Moments for Green Channel")
    Mean2 = np.mean(G);
    print (Mean2)
    Variance2 = np.var(G)
    Variance2 = math.sqrt(Variance2)
    print(Variance2)
    Skewness2=skew(G.reshape(-1))
    print (Skewness2)
    Kurtosis2=kurtosis(G.reshape(-1))
    print (Kurtosis2)
    entropy2 = skimage.measure.shannon_entropy(G)
    print (entropy2)
    G = skimage.img_as_ubyte(G)
    g2 = skimage.feature.greycomatrix(G, [1], [0], levels=256, symmetric=False, normed=True)
    Cont2 =skimage.feature.greycoprops(g2, 'contrast')[0][0]
    print (Cont2)
    Energ2 =skimage.feature.greycoprops(g2, 'energy')[0][0]
    print (Energ2)
    Homo2=skimage.feature.greycoprops(g2, 'homogeneity')[0][0]
    print (Homo2)
    Corre2=skimage.feature.greycoprops(g2, 'correlation')[0][0]
    print (Corre2)
    
    #  Feature Selection from Blue Channel
    print ("Calucation of First Four Statistical Four Moments for Blue Channel")
    Mean3 = np.mean(B);
    print (Mean3)
    Variance3 = np.var(B)
    Variance3 = math.sqrt(Variance3)
    print(Variance3)
    Skewness3=skew(B.reshape(-1))
    print (Skewness3)
    Kurtosis3=kurtosis(B.reshape(-1))
    print (Kurtosis3)
    entropy3 = skimage.measure.shannon_entropy(B)
    print (entropy3)
    B = skimage.img_as_ubyte(B)
    g3 = skimage.feature.greycomatrix(B, [1], [0], levels=256, symmetric=False, normed=True)
    Cont3 =skimage.feature.greycoprops(g3, 'contrast')[0][0]
    print (Cont3)
    Energ3 =skimage.feature.greycoprops(g3, 'energy')[0][0]
    print (Energ3)
    Homo3=skimage.feature.greycoprops(g3, 'homogeneity')[0][0]
    print (Homo3)
    Corre3=skimage.feature.greycoprops(g3, 'correlation')[0][0]
    print (Corre3)
    

    print("How many Times" )
    print (i)
    
    # Create File
    outWorkbook = xlsxwriter.Workbook("out.xlsx")
    outSheet = outWorkbook.add_worksheet()
    
    # Declare Data
    values = [Mean1, Variance1, Skewness1, Kurtosis1,entropy1, Cont1,Energ1, Homo1, Corre1, Mean2, Variance2, Skewness2, Kurtosis2, entropy2, Cont2, Energ2, Homo2, Corre2, Mean3, Variance3, Skewness3, Kurtosis3, entropy3, Cont3, Energ3, Homo3, Corre3 ]
    
    
   
    # Write data to file
    outWorkbook = openpyxl.load_workbook("out.xlsx") 
    outSheet = outWorkbook.active
    outSheet.append(values)

    outWorkbook.save(filename="out.xlsx")

df = pd.read_excel("out.xlsx")
df.to_csv("data.csv")

# Importing the dataset
dataset = pd.read_csv('data.csv')
dataset1 = pd.read_csv('Labels1.csv')
X = dataset.iloc[:, 0:27].values
y = dataset1.iloc[:, 0].values
    
# Training and Testing Data (divide the data into two part)
from sklearn.model_selection import train_test_split
X_train, X_test, y_train, y_test =train_test_split(X,y,train_size=0.75, random_state=0)

from sklearn.preprocessing import StandardScaler
sc = StandardScaler()
X_train = sc.fit_transform(X_train)
X_test = sc.transform(X_test)

from sklearn.tree import DecisionTreeClassifier
classifier=DecisionTreeClassifier(criterion='entropy', random_state=0)
classifier.fit(X_train,y_train)

y_pred= classifier.predict(X_test)

from sklearn.metrics import confusion_matrix
cm = confusion_matrix(y_test,y_pred)

from sklearn.metrics import accuracy_score
accuracy_score(y_test,y_pred)

# Making the Confusion Matrix
cm = confusion_matrix(y_pred, y_test)
sns.heatmap(cm,annot=True)
pyplot.savefig('h.png')
print(cm)

from sklearn.metrics import accuracy_score
accuracy_score(y_test,y_pred)

from sklearn.metrics import classification_report
print(classification_report(y_test,y_pred))

import pickle 
  
# Save the trained model as a pickle string. 
with open ('classifier_pickle','wb') as f:
    pickle.dump(classifier,f)
